# -*- coding: utf-8 -*-
"""
Chuan hoa anh minh hoa khach san do NGUOI DUNG tu tai ve (khong tai ho, chi xu ly).

Cach dung:
    1) Dat cac file anh goc (bat ky kich thuoc/dinh dang jpg, jpeg, png, webp)
       vao thu muc assets/hotel_photos_raw/
    2) Chay:  python prepare_photos.py
    3) Script se:
         - Resize + nen lai tung anh (khong can anh dung kich thuoc san)
         - Doi ten gon: beach_01.jpg, beach_02.jpg, ...
         - Luu vao assets/hotel_photos/
         - Tu dong cap nhat cot Ten_file trong assets/hotel_photos/photos.xlsx
           (giu nguyen Nguon_anh / Ghi_chu ban da dien tu truoc, neu co)

YEU CAU ANH DAU VAO (de bam khi tu chuan bi):
    - Dinh dang : JPG, JPEG, PNG hoac WEBP
    - Kich thuoc : khong quan trong — script tu resize/crop vuong, khuyen nghi
                   canh ngan >= 500px de anh khong bi mo sau khi resize
    - Dung luong : moi anh nen duoi ~5MB truoc khi xu ly (de script chay nhanh)
    - So luong   : bao nhieu cung duoc, khong bat buoc dung 30 anh
"""
import os
import sys

RAW_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "hotel_photos_raw")
OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "hotel_photos")
XLSX_PATH = os.path.join(OUT_DIR, "photos.xlsx")

TARGET_SIZE = 640          # anh vuong 640x640 — du net cho the nho, nhe cho web
JPEG_QUALITY = 82
VALID_EXT = (".jpg", ".jpeg", ".png", ".webp")


def log(m):
    print(m, flush=True)


def resize_and_crop_square(im, size):
    """Cat vuong chinh giua roi resize — dung ImageOps.fit de khong bi meo anh."""
    from PIL import ImageOps
    return ImageOps.fit(im, (size, size), method=3)  # 3 = Image.LANCZOS


def process_images():
    from PIL import Image

    if not os.path.isdir(RAW_DIR):
        os.makedirs(RAW_DIR, exist_ok=True)
        sys.exit(
            f"[LOI] Chua co thu muc {RAW_DIR}\n"
            f"      Da tu tao thu muc nay giup ban — hay bo anh goc vao do roi chay lai script."
        )

    files = sorted(
        f for f in os.listdir(RAW_DIR)
        if f.lower().endswith(VALID_EXT) and not f.startswith(".")
    )
    if not files:
        sys.exit(f"[LOI] Khong tim thay anh nao trong {RAW_DIR} (dinh dang: {VALID_EXT})")

    os.makedirs(OUT_DIR, exist_ok=True)
    saved_names = []
    for i, fname in enumerate(files, start=1):
        src_path = os.path.join(RAW_DIR, fname)
        try:
            im = Image.open(src_path).convert("RGB")
        except Exception as exc:
            log(f"[bo qua] {fname}: khong doc duoc anh ({exc})")
            continue

        im2 = resize_and_crop_square(im, TARGET_SIZE)
        out_name = f"beach_{i:02d}.jpg"
        out_path = os.path.join(OUT_DIR, out_name)
        im2.save(out_path, "JPEG", quality=JPEG_QUALITY, optimize=True)
        size_kb = round(os.path.getsize(out_path) / 1024, 1)
        log(f"[anh] {fname}  ->  {out_name}  ({TARGET_SIZE}x{TARGET_SIZE}, {size_kb} KB)")
        saved_names.append(out_name)

    return saved_names


def update_excel(saved_names):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    FONT_NAME = "Arial"
    HEADER_ROW = 8

    # Doc lai Nguon_anh / Ghi_chu cu (theo ten file) neu file xlsx da co san, de khong mat du lieu ban da dien
    old_notes = {}
    if os.path.exists(XLSX_PATH):
        try:
            wb_old = openpyxl.load_workbook(XLSX_PATH, data_only=True)
            ws_old = wb_old["Anh_minh_hoa"]
            for row in ws_old.iter_rows(min_row=HEADER_ROW + 1, values_only=False):
                fname = row[1].value  # cot B = Ten_file
                if fname:
                    old_notes[str(fname).strip()] = {
                        "su_dung": row[2].value if row[2].value not in (None, "") else "TRUE",
                        "nguon": row[3].value or "",
                        "ghi_chu": row[4].value or "",
                    }
        except Exception as exc:
            log(f"[canh bao] khong doc duoc photos.xlsx cu, se tao moi ({exc})")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Anh_minh_hoa"

    header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="0E7C86")
    edit_fill = PatternFill("solid", fgColor="FFF6D9")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws["A1"] = "HƯỚNG DẪN — Danh sách ảnh minh hoạ khách sạn"
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=13, color="073B4C")
    ws.merge_cells("A1:E1")

    legend_lines = [
        "Danh sach nay duoc prepare_photos.py tu dong cap nhat cot Ten_file moi lan chay.",
        "Ban chi can dien 2 cot mau vang: Nguon_anh (de trich dan) va Ghi_chu (neu muon).",
        "Cot Su_dung: TRUE = dung anh nay, FALSE = tam an anh (khong can xoa file).",
    ]
    for i, line in enumerate(legend_lines, start=2):
        ws.cell(row=i, column=1, value=line)
        ws.cell(row=i, column=1).font = Font(name=FONT_NAME, italic=True, size=10, color="444444")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)

    headers = ["STT", "Ten_file (tên file ảnh)", "Su_dung (TRUE/FALSE)", "Nguon_anh (để trích dẫn)", "Ghi_chu"]
    widths = [7, 30, 20, 32, 30]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=HEADER_ROW, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    for i, fname in enumerate(saved_names, start=1):
        r = HEADER_ROW + i
        prev = old_notes.get(fname, {})
        vals = [i, fname, prev.get("su_dung", "TRUE"), prev.get("nguon", ""), prev.get("ghi_chu", "")]
        for col, val in enumerate(vals, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = border
            c.font = Font(name=FONT_NAME, size=10)
            if col in (2, 4, 5):
                c.fill = edit_fill
                c.alignment = wrap
            elif col == 3:
                c.fill = edit_fill
                c.alignment = Alignment(horizontal="center")
            else:
                c.alignment = Alignment(horizontal="center")

    ws.freeze_panes = f"A{HEADER_ROW + 1}"
    ws.row_dimensions[1].height = 22
    wb.save(XLSX_PATH)
    log(f"[excel] da cap nhat {XLSX_PATH} ({len(saved_names)} anh)")


if __name__ == "__main__":
    names = process_images()
    update_excel(names)
    log(f"\nHoan tat {len(names)} anh. Mo assets/hotel_photos/photos.xlsx de dien Nguon_anh/Ghi_chu neu muon.")
    log("Chay lai: streamlit run app.py")
